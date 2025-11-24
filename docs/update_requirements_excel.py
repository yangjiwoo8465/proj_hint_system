"""
기존 요구사항 정의서에 개인화 힌트 시스템 추가
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import shutil
import os

# 기존 파일 경로
source_file = 'C:/Users/playdata2/Desktop/playdata/Workspace/팀프로젝트5/0.기획_요구사항 정의서_v.0.2.xlsx'
output_file = 'C:/Users/playdata2/Desktop/playdata/Workspace/팀프로젝트5/5th-project_mvp/docs/0.기획_요구사항 정의서_v.0.3_개인화힌트추가.xlsx'

# 파일 복사
shutil.copy(source_file, output_file)

# 워크북 로드
wb = load_workbook(output_file)

# 스타일 정의
header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
subheader_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
white_font = Font(color="FFFFFF", bold=True, size=12)
bold_font = Font(bold=True, size=11)
normal_font = Font(size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def set_cell_style(cell, font, fill=None, alignment=None, border_style=border):
    cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    else:
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    cell.border = border_style

# ====================
# 새 시트 추가: 개인화 힌트 시스템
# ====================
ws_hint = wb.create_sheet("5. 개인화 힌트 시스템", 4)  # 5번째 위치에 삽입
ws_hint.column_dimensions['A'].width = 25
ws_hint.column_dimensions['B'].width = 75

# 제목
row = 1
ws_hint[f'A{row}'] = "개인화 힌트 시스템 (Personalized Hint System)"
ws_hint.merge_cells(f'A{row}:B{row}')
set_cell_style(ws_hint[f'A{row}'], Font(bold=True, size=14, color="FFFFFF"),
               PatternFill(start_color="764BA2", end_color="764BA2", fill_type="solid"),
               Alignment(horizontal='center', vertical='center'))
ws_hint.row_dimensions[row].height = 30
row += 2

# 5.1 개요
ws_hint[f'A{row}'] = "5.1 개요"
ws_hint.merge_cells(f'A{row}:B{row}')
set_cell_style(ws_hint[f'A{row}'], Font(bold=True, size=12), subheader_fill)
row += 1

ws_hint[f'A{row}'] = "목적"
ws_hint[f'B{row}'] = "사용자의 실력 수준에 따라 개인화된 힌트를 제공하여 효과적인 학습 경험 제공"
set_cell_style(ws_hint[f'A{row}'], bold_font)
set_cell_style(ws_hint[f'B{row}'], normal_font)
row += 1

ws_hint[f'A{row}'] = "핵심 기능"
ws_hint[f'B{row}'] = ("• 사용자 행동 추적 및 실력 점수 자동 계산\n"
                      "• 3단계 힌트 레벨 시스템 (기초/보통/실력자)\n"
                      "• 관리자 패널을 통한 사용자 실력 지표 관리\n"
                      "• 자동/수동 모드 전환")
set_cell_style(ws_hint[f'A{row}'], bold_font)
set_cell_style(ws_hint[f'B{row}'], normal_font)
ws_hint.row_dimensions[row].height = 60
row += 2

# 5.2 실력 지표 관리
ws_hint[f'A{row}'] = "5.2 실력 지표 관리"
ws_hint.merge_cells(f'A{row}:B{row}')
set_cell_style(ws_hint[f'A{row}'], Font(bold=True, size=12), subheader_fill)
row += 1

# 실력 점수
ws_hint[f'A{row}'] = "실력 점수 (Skill Score)"
ws_hint[f'B{row}'] = ("범위: 0~100 (Float)\n"
                      "의미: 높을수록 초보, 낮을수록 실력자\n"
                      "기본값: 50.0\n"
                      "계산: 최근 10개 문제 풀이 세션의 난이도 점수 평균")
set_cell_style(ws_hint[f'A{row}'], bold_font)
set_cell_style(ws_hint[f'B{row}'], normal_font)
ws_hint.row_dimensions[row].height = 60
row += 1

# 실력 모드
ws_hint[f'A{row}'] = "실력 모드 (Skill Mode)"
ws_hint[f'B{row}'] = ("자동 모드 (auto):\n"
                      "  - 문제를 풀 때마다 실력 점수가 자동으로 업데이트\n"
                      "  - 실력 점수에 따라 힌트 레벨이 자동 조정\n\n"
                      "수동 모드 (manual):\n"
                      "  - 관리자가 설정한 값으로 고정\n"
                      "  - 문제를 풀어도 자동 업데이트되지 않음")
set_cell_style(ws_hint[f'A{row}'], bold_font)
set_cell_style(ws_hint[f'B{row}'], normal_font)
ws_hint.row_dimensions[row].height = 90
row += 1

# 힌트 레벨
ws_hint[f'A{row}'] = "힌트 레벨 (Hint Level)"
ws_hint[f'B{row}'] = ("레벨 1 (기초): 실력 점수 70 이상 → 구체적 힌트 (함수명 직접 언급)\n"
                      "레벨 2 (보통): 실력 점수 40~70 → 개념 힌트\n"
                      "레벨 3 (실력자): 실력 점수 40 미만 → 소크라테스식 질문")
set_cell_style(ws_hint[f'A{row}'], bold_font)
set_cell_style(ws_hint[f'B{row}'], normal_font)
ws_hint.row_dimensions[row].height = 60
row += 2

# 5.3 난이도 점수 계산 알고리즘
ws_hint[f'A{row}'] = "5.3 난이도 점수 계산 알고리즘"
ws_hint.merge_cells(f'A{row}:B{row}')
set_cell_style(ws_hint[f'A{row}'], Font(bold=True, size=12), subheader_fill)
row += 1

ws_hint[f'A{row}'] = "계산 공식"
ws_hint[f'B{row}'] = ("총점 = 0~100 (높을수록 어려워함)\n\n"
                      "1. 힌트 요청 시간 점수:\n"
                      "   • 1분 이내: +30점\n"
                      "   • 5분 이내: +20점\n"
                      "   • 10분 이내: +10점\n\n"
                      "2. 힌트 요청 횟수: 횟수 × 10 (최대 30점)\n\n"
                      "3. 코드 실행 횟수: 횟수 × 3 (최대 20점)\n\n"
                      "4. 코드 길이 점수:\n"
                      "   • 50자 미만: +20점\n"
                      "   • 100자 미만: +10점")
set_cell_style(ws_hint[f'A{row}'], bold_font)
set_cell_style(ws_hint[f'B{row}'], normal_font)
ws_hint.row_dimensions[row].height = 180
row += 2

# 5.4 레벨별 힌트 스타일
ws_hint[f'A{row}'] = "5.4 레벨별 힌트 스타일"
ws_hint.merge_cells(f'A{row}:B{row}')
set_cell_style(ws_hint[f'A{row}'], Font(bold=True, size=12), subheader_fill)
row += 1

# 레벨 1
ws_hint[f'A{row}'] = "레벨 1 (기초 - 구체적 힌트)"
ws_hint[f'B{row}'] = ("대상: 실력 점수 70 이상 (초보자)\n\n"
                      "특징:\n"
                      "  - 필요한 함수명, 라이브러리, 메서드를 직접 언급\n"
                      "  - 단계별 구체적인 코드 작성 방법 제시\n"
                      "  - 150자 이내로 간단명료하게 제공\n\n"
                      "예시:\n"
                      '  "N줄을 입력받으려면 for _ in range(N)을 사용하고,\n'
                      '   각 줄을 list()로 변환해서 board에 append() 하세요."')
set_cell_style(ws_hint[f'A{row}'], bold_font)
set_cell_style(ws_hint[f'B{row}'], normal_font)
ws_hint.row_dimensions[row].height = 120
row += 1

# 레벨 2
ws_hint[f'A{row}'] = "레벨 2 (보통 - 개념 힌트)"
ws_hint[f'B{row}'] = ("대상: 실력 점수 40~70 (중급자)\n\n"
                      "특징:\n"
                      "  - 함수명을 직접 언급하지 않고 개념으로 유도\n"
                      "  - 필요한 자료구조나 알고리즘 개념 안내\n"
                      "  - 180자 이내로 제공\n\n"
                      "예시:\n"
                      '  "N줄의 보드를 저장하려면 2차원 리스트가 필요합니다.\n'
                      '   반복문으로 각 줄을 입력받아 추가하세요."')
set_cell_style(ws_hint[f'A{row}'], bold_font)
set_cell_style(ws_hint[f'B{row}'], normal_font)
ws_hint.row_dimensions[row].height = 120
row += 1

# 레벨 3
ws_hint[f'A{row}'] = "레벨 3 (실력자 - 소크라테스식)"
ws_hint[f'B{row}'] = ("대상: 실력 점수 40 미만 (고급자)\n\n"
                      "특징:\n"
                      "  - 질문 형식으로만 힌트 제공\n"
                      "  - 학생이 스스로 답을 찾도록 유도\n"
                      "  - 한 번에 하나의 질문만 (200자 이내)\n"
                      "  - 평가나 분석 내용 포함하지 않음\n\n"
                      "예시:\n"
                      '  "전체 보드의 상태를 어떻게 입력받아 저장할 수 있을까요?"')
set_cell_style(ws_hint[f'A{row}'], bold_font)
set_cell_style(ws_hint[f'B{row}'], normal_font)
ws_hint.row_dimensions[row].height = 140
row += 2

# 5.5 관리자 기능
ws_hint[f'A{row}'] = "5.5 관리자 기능"
ws_hint.merge_cells(f'A{row}:B{row}')
set_cell_style(ws_hint[f'A{row}'], Font(bold=True, size=12), subheader_fill)
row += 1

ws_hint[f'A{row}'] = "사용자 실력 지표 관리"
ws_hint[f'B{row}'] = ("• 모든 사용자 목록 조회 (실력 점수, 실력 모드, 힌트 레벨 포함)\n"
                      "• 실력 지표 수정 기능:\n"
                      "  - 실력 모드: 자동 ↔ 수동 전환\n"
                      "  - 실력 점수: 수동 모드일 때만 수정 가능\n"
                      "  - 힌트 레벨: 수동 모드일 때만 수정 가능\n"
                      "• 인라인 편집으로 빠른 수정\n"
                      "• 자동 모드일 때 수정 불가 필드는 비활성화 표시")
set_cell_style(ws_hint[f'A{row}'], bold_font)
set_cell_style(ws_hint[f'B{row}'], normal_font)
ws_hint.row_dimensions[row].height = 100
row += 2

# ====================
# 새 시트 추가: 데이터 모델 (개인화 힌트)
# ====================
ws_model = wb.create_sheet("6. 데이터 모델 (힌트)", 5)
ws_model.column_dimensions['A'].width = 20
ws_model.column_dimensions['B'].width = 20
ws_model.column_dimensions['C'].width = 15
ws_model.column_dimensions['D'].width = 50

# 제목
row = 1
ws_model[f'A{row}'] = "개인화 힌트 시스템 - 데이터 모델"
ws_model.merge_cells(f'A{row}:D{row}')
set_cell_style(ws_model[f'A{row}'], Font(bold=True, size=14, color="FFFFFF"),
               PatternFill(start_color="764BA2", end_color="764BA2", fill_type="solid"),
               Alignment(horizontal='center', vertical='center'))
ws_model.row_dimensions[row].height = 30
row += 2

# User 모델 확장
ws_model[f'A{row}'] = "User 모델 (확장)"
ws_model.merge_cells(f'A{row}:D{row}')
set_cell_style(ws_model[f'A{row}'], white_font, header_fill, Alignment(horizontal='center'))
row += 1

# 헤더
headers = ["필드명", "타입", "기본값", "설명"]
for col, header in enumerate(headers, 1):
    cell = ws_model.cell(row, col, header)
    set_cell_style(cell, bold_font, subheader_fill, Alignment(horizontal='center'))
row += 1

# User 필드
user_fields = [
    ("skill_score", "FloatField", "50.0", "실력 점수 (0~100, 높을수록 초보)"),
    ("skill_mode", "CharField", "auto", "실력 지표 모드 (auto/manual)"),
    ("hint_level", "IntegerField", "2", "힌트 레벨 (1:기초, 2:보통, 3:실력자)"),
]

for field_name, field_type, default, desc in user_fields:
    ws_model[f'A{row}'] = field_name
    ws_model[f'B{row}'] = field_type
    ws_model[f'C{row}'] = default
    ws_model[f'D{row}'] = desc
    for col in ['A', 'B', 'C', 'D']:
        set_cell_style(ws_model[f'{col}{row}'], normal_font)
    row += 1

row += 1

# ProblemSession 모델
ws_model[f'A{row}'] = "ProblemSession 모델 (신규)"
ws_model.merge_cells(f'A{row}:D{row}')
set_cell_style(ws_model[f'A{row}'], white_font, header_fill, Alignment(horizontal='center'))
row += 1

# 헤더
for col, header in enumerate(headers, 1):
    cell = ws_model.cell(row, col, header)
    set_cell_style(cell, bold_font, subheader_fill, Alignment(horizontal='center'))
row += 1

# ProblemSession 필드
session_fields = [
    ("user", "ForeignKey", "-", "사용자 (User 모델 참조)"),
    ("problem", "ForeignKey", "-", "문제 (Problem 모델 참조)"),
    ("started_at", "DateTimeField", "auto_now_add", "세션 시작 시간"),
    ("first_hint_at", "DateTimeField", "null", "첫 힌트 요청 시간"),
    ("first_run_at", "DateTimeField", "null", "첫 코드 실행 시간"),
    ("solved_at", "DateTimeField", "null", "문제 해결 시간"),
    ("hint_count", "IntegerField", "0", "힌트 요청 횟수"),
    ("run_count", "IntegerField", "0", "코드 실행 횟수"),
    ("max_code_length", "IntegerField", "0", "최대 코드 길이"),
    ("is_solved", "BooleanField", "False", "해결 여부"),
]

for field_name, field_type, default, desc in session_fields:
    ws_model[f'A{row}'] = field_name
    ws_model[f'B{row}'] = field_type
    ws_model[f'C{row}'] = default
    ws_model[f'D{row}'] = desc
    for col in ['A', 'B', 'C', 'D']:
        set_cell_style(ws_model[f'{col}{row}'], normal_font)
    row += 1

row += 2

# 제약사항
ws_model[f'A{row}'] = "제약사항"
ws_model[f'B{row}'] = "unique_together = ['user', 'problem']"
ws_model.merge_cells(f'B{row}:D{row}')
set_cell_style(ws_model[f'A{row}'], bold_font, subheader_fill)
set_cell_style(ws_model[f'B{row}'], normal_font)

# ====================
# 새 시트 추가: API 명세 (개인화 힌트)
# ====================
ws_api = wb.create_sheet("7. API 명세 (힌트)", 6)
ws_api.column_dimensions['A'].width = 15
ws_api.column_dimensions['B'].width = 40
ws_api.column_dimensions['C'].width = 15
ws_api.column_dimensions['D'].width = 35

# 제목
row = 1
ws_api[f'A{row}'] = "개인화 힌트 시스템 - API 명세"
ws_api.merge_cells(f'A{row}:D{row}')
set_cell_style(ws_api[f'A{row}'], Font(bold=True, size=14, color="FFFFFF"),
               PatternFill(start_color="764BA2", end_color="764BA2", fill_type="solid"),
               Alignment(horizontal='center', vertical='center'))
ws_api.row_dimensions[row].height = 30
row += 2

# 헤더
headers = ["구분", "엔드포인트", "메서드", "설명"]
for col, header in enumerate(headers, 1):
    cell = ws_api.cell(row, col, header)
    set_cell_style(cell, white_font, header_fill, Alignment(horizontal='center'))
row += 1

# API 목록
apis = [
    ("힌트 요청", "/api/v1/coding-test/hint/", "POST", "사용자 힌트 레벨에 따른 AI 힌트 생성"),
    ("사용자 목록", "/api/v1/admin/users/", "GET", "모든 사용자 목록 조회 (실력 지표 포함)"),
    ("실력 지표 수정", "/api/v1/admin/users/{id}/skill/", "PATCH", "사용자 실력 지표 업데이트"),
    ("사용자 상태", "/api/v1/admin/users/{id}/", "PATCH", "사용자 활성화 상태 변경"),
    ("사용자 삭제", "/api/v1/admin/users/{id}/delete/", "DELETE", "사용자 삭제"),
]

for category, endpoint, method, desc in apis:
    ws_api[f'A{row}'] = category
    ws_api[f'B{row}'] = endpoint
    ws_api[f'C{row}'] = method
    ws_api[f'D{row}'] = desc
    for col in ['A', 'B', 'C', 'D']:
        set_cell_style(ws_api[f'{col}{row}'], normal_font)
    ws_api.row_dimensions[row].height = 30
    row += 1

row += 2

# API 상세 - 힌트 요청
ws_api[f'A{row}'] = "힌트 요청 API 상세"
ws_api.merge_cells(f'A{row}:D{row}')
set_cell_style(ws_api[f'A{row}'], bold_font, subheader_fill)
row += 1

ws_api[f'A{row}'] = "Request Body"
ws_api[f'B{row}'] = '{\n  "problem_id": "string",\n  "user_code": "string"\n}'
ws_api.merge_cells(f'B{row}:D{row}')
set_cell_style(ws_api[f'A{row}'], bold_font)
set_cell_style(ws_api[f'B{row}'], normal_font)
ws_api.row_dimensions[row].height = 40
row += 1

ws_api[f'A{row}'] = "Response"
ws_api[f'B{row}'] = '{\n  "success": true,\n  "data": {\n    "hint": "string",\n    "problem_id": "string"\n  }\n}'
ws_api.merge_cells(f'B{row}:D{row}')
set_cell_style(ws_api[f'A{row}'], bold_font)
set_cell_style(ws_api[f'B{row}'], normal_font)
ws_api.row_dimensions[row].height = 60
row += 1

ws_api[f'A{row}'] = "처리 흐름"
ws_api[f'B{row}'] = ("1. request.user.hint_level 조회\n"
                     "2. hint_level에 따라 다른 프롬프트 생성\n"
                     "3. Hugging Face API 호출\n"
                     "4. HintRequest 기록 저장\n"
                     "5. 힌트 반환")
ws_api.merge_cells(f'B{row}:D{row}')
set_cell_style(ws_api[f'A{row}'], bold_font)
set_cell_style(ws_api[f'B{row}'], normal_font)
ws_api.row_dimensions[row].height = 70
row += 2

# API 상세 - 실력 지표 수정
ws_api[f'A{row}'] = "실력 지표 수정 API 상세"
ws_api.merge_cells(f'A{row}:D{row}')
set_cell_style(ws_api[f'A{row}'], bold_font, subheader_fill)
row += 1

ws_api[f'A{row}'] = "Request Body"
ws_api[f'B{row}'] = '{\n  "skill_score": 75.5,\n  "skill_mode": "manual",\n  "hint_level": 1\n}'
ws_api.merge_cells(f'B{row}:D{row}')
set_cell_style(ws_api[f'A{row}'], bold_font)
set_cell_style(ws_api[f'B{row}'], normal_font)
ws_api.row_dimensions[row].height = 50
row += 1

ws_api[f'A{row}'] = "Validation"
ws_api[f'B{row}'] = ("• skill_mode: 'auto' | 'manual'만 허용\n"
                     "• hint_level: 1|2|3만 허용\n"
                     "• skill_score: 0~100 범위")
ws_api.merge_cells(f'B{row}:D{row}')
set_cell_style(ws_api[f'A{row}'], bold_font)
set_cell_style(ws_api[f'B{row}'], normal_font)
ws_api.row_dimensions[row].height = 50

# ====================
# 새 시트 추가: 시스템 플로우
# ====================
ws_flow = wb.create_sheet("8. 시스템 플로우 (힌트)", 7)
ws_flow.column_dimensions['A'].width = 30
ws_flow.column_dimensions['B'].width = 70

# 제목
row = 1
ws_flow[f'A{row}'] = "개인화 힌트 시스템 - 플로우"
ws_flow.merge_cells(f'A{row}:B{row}')
set_cell_style(ws_flow[f'A{row}'], Font(bold=True, size=14, color="FFFFFF"),
               PatternFill(start_color="764BA2", end_color="764BA2", fill_type="solid"),
               Alignment(horizontal='center', vertical='center'))
ws_flow.row_dimensions[row].height = 30
row += 2

# 자동 모드 플로우
ws_flow[f'A{row}'] = "자동 모드 - 실력 지표 갱신 흐름"
ws_flow.merge_cells(f'A{row}:B{row}')
set_cell_style(ws_flow[f'A{row}'], bold_font, subheader_fill)
row += 1

flow_auto = """1. 사용자가 문제 풀이 시작
   ↓
2. ProblemSession 생성 (started_at 기록)
   ↓
3. 힌트 요청 시
   - first_hint_at 기록 (첫 요청인 경우)
   - hint_count 증가
   ↓
4. 코드 실행 시
   - first_run_at 기록 (첫 실행인 경우)
   - run_count 증가
   - max_code_length 업데이트
   ↓
5. 문제 해결 시
   - solved_at 기록
   - is_solved = True
   - user.update_skill_score() 호출
   ↓
6. update_skill_score() 메서드
   - skill_mode가 'auto'인지 확인
   - 최근 10개 세션의 난이도 점수 계산
   - 평균 점수로 skill_score 업데이트
   - 점수에 따라 hint_level 자동 조정
   ↓
7. 다음 힌트 요청 시 새로운 hint_level 적용"""

ws_flow[f'A{row}'] = flow_auto
ws_flow.merge_cells(f'A{row}:B{row}')
set_cell_style(ws_flow[f'A{row}'], normal_font)
ws_flow.row_dimensions[row].height = 300
row += 2

# 수동 모드 플로우
ws_flow[f'A{row}'] = "수동 모드 - 관리자 설정 흐름"
ws_flow.merge_cells(f'A{row}:B{row}')
set_cell_style(ws_flow[f'A{row}'], bold_font, subheader_fill)
row += 1

flow_manual = """1. 관리자가 "관리자 패널" 접속
   ↓
2. "사용자 관리" 탭 선택
   ↓
3. 특정 사용자의 "실력 수정" 클릭
   ↓
4. 편집 모드 진입
   - skill_mode를 'manual'로 선택
   - skill_score 직접 입력 (예: 75.5)
   - hint_level 직접 선택 (예: 레벨 1)
   ↓
5. "저장" 버튼 클릭
   ↓
6. API 호출: PATCH /api/v1/admin/users/{id}/skill/
   ↓
7. 백엔드: User 모델 업데이트
   - skill_score = 75.5
   - skill_mode = 'manual'
   - hint_level = 1
   ↓
8. 사용자는 다음 힌트 요청부터 레벨 1 힌트 받음
   (문제를 풀어도 자동 업데이트되지 않음)"""

ws_flow[f'A{row}'] = flow_manual
ws_flow.merge_cells(f'A{row}:B{row}')
set_cell_style(ws_flow[f'A{row}'], normal_font)
ws_flow.row_dimensions[row].height = 250

# 파일 저장
wb.save(output_file)
print("Updated requirements file created:", output_file)
print("Added sheets:")
print("  - 5. 개인화 힌트 시스템")
print("  - 6. 데이터 모델 (힌트)")
print("  - 7. API 명세 (힌트)")
print("  - 8. 시스템 플로우 (힌트)")
