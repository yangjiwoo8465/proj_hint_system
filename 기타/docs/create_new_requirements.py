"""
개인화 힌트 시스템 요구사항 정의서 (새로 작성)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 워크북 생성
wb = Workbook()
wb.remove(wb.active)  # 기본 시트 제거

# 스타일 정의
header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
subheader_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
purple_fill = PatternFill(start_color="764BA2", end_color="764BA2", fill_type="solid")
white_font = Font(color="FFFFFF", bold=True, size=12)
bold_font = Font(bold=True, size=11)
normal_font = Font(size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def set_cell_style(cell, font, fill=None, alignment=None, border_style=border):
    cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    else:
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    cell.border = border_style

# ====================
# 1. 프로젝트 개요
# ====================
ws1 = wb.create_sheet("1.프로젝트개요")
ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 80

# 제목
row = 1
ws1[f'A{row}'] = "개인화된 힌트 시스템 - 요구사항 정의서"
ws1.merge_cells(f'A{row}:B{row}')
set_cell_style(ws1[f'A{row}'], Font(bold=True, size=16, color="FFFFFF"),
               purple_fill, Alignment(horizontal='center', vertical='center'))
ws1.row_dimensions[row].height = 35
row += 2

data = [
    ("프로젝트명", "개인화된 힌트 시스템 (Personalized Hint System)"),
    ("작성일", "2025-01-06"),
    ("버전", "v1.0"),
    ("작성자", "개발팀"),
    ("프로젝트 목적", "사용자의 실력 수준에 맞는 개인화된 힌트를 제공하여 효과적인 학습을 지원하는 시스템"),
    ("주요 기능", "• 사용자 행동 추적 및 실력 점수 자동 계산\n• 3단계 힌트 레벨 시스템 (기초/보통/실력자)\n• 실력 점수 기반 자동 힌트 레벨 조정\n• 관리자 패널을 통한 사용자 실력 지표 관리\n• 자동/수동 모드 전환 기능"),
    ("적용 범위", "• 코딩 테스트 플랫폼의 힌트 시스템\n• 관리자 패널의 사용자 관리 기능"),
    ("기술 스택", "• Frontend: React 18, Redux, Monaco Editor, React Router\n• Backend: Django 4.2, Django REST Framework\n• Database: PostgreSQL 15\n• AI: Hugging Face Inference Providers (Qwen 7B/32B)\n• Infrastructure: Docker, Docker Compose"),
]

for label, value in data:
    ws1[f'A{row}'] = label
    ws1[f'B{row}'] = value
    set_cell_style(ws1[f'A{row}'], bold_font, subheader_fill)
    set_cell_style(ws1[f'B{row}'], normal_font)

    if label in ["주요 기능", "기술 스택", "적용 범위"]:
        ws1.row_dimensions[row].height = 80
    else:
        ws1.row_dimensions[row].height = 25
    row += 1

# ====================
# 2. 기능 요구사항
# ====================
ws2 = wb.create_sheet("2.기능요구사항")
ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 25
ws2.column_dimensions['C'].width = 50
ws2.column_dimensions['D'].width = 15

# 제목
row = 1
ws2[f'A{row}'] = "기능 요구사항"
ws2.merge_cells(f'A{row}:D{row}')
set_cell_style(ws2[f'A{row}'], Font(bold=True, size=14, color="FFFFFF"),
               purple_fill, Alignment(horizontal='center', vertical='center'))
ws2.row_dimensions[row].height = 30
row += 2

# 헤더
headers = ["ID", "기능명", "설명", "우선순위"]
for col, header in enumerate(headers, 1):
    cell = ws2.cell(row, col, header)
    set_cell_style(cell, white_font, header_fill, Alignment(horizontal='center', vertical='center'))
row += 1

# 요구사항 데이터
requirements = [
    ("FR-001", "사용자 행동 추적", "문제 선택 시점부터 첫 힌트 요청까지의 시간, 첫 실행까지의 시간, 실행 횟수, 코드 길이를 자동으로 추적하여 ProblemSession에 기록", "높음"),
    ("FR-002", "실력 점수 자동 계산", "사용자의 문제 풀이 행동을 기반으로 0-100 범위의 실력 점수를 자동 계산 (높을수록 초보)", "높음"),
    ("FR-003", "난이도 점수 계산", "힌트 요청 시간, 힌트 횟수, 실행 횟수, 코드 길이를 종합하여 세션별 난이도 점수 산출", "높음"),
    ("FR-004", "힌트 레벨 자동 조정", "실력 점수에 따라 힌트 레벨을 자동 조정\n- 70점 이상: 레벨 1 (기초)\n- 40-70점: 레벨 2 (보통)\n- 40점 미만: 레벨 3 (실력자)", "높음"),
    ("FR-005", "레벨별 맞춤 힌트 제공", "사용자의 힌트 레벨에 따라 서로 다른 스타일의 힌트를 AI가 생성하여 제공", "높음"),
    ("FR-006", "자동 모드", "문제를 풀 때마다 실력 점수와 힌트 레벨이 자동으로 업데이트되는 모드", "중간"),
    ("FR-007", "수동 모드", "관리자가 설정한 실력 점수와 힌트 레벨이 고정되어 자동 업데이트되지 않는 모드", "중간"),
    ("FR-008", "관리자 사용자 목록 조회", "관리자 패널에서 모든 사용자의 실력 지표(점수, 모드, 레벨)를 조회", "중간"),
    ("FR-009", "관리자 실력 지표 수정", "관리자가 특정 사용자의 실력 점수, 실력 모드, 힌트 레벨을 인라인 편집으로 수정", "중간"),
    ("FR-010", "사용자 상태 관리", "관리자가 사용자 계정을 활성화/비활성화하거나 삭제", "낮음"),
]

for req_id, name, desc, priority in requirements:
    ws2[f'A{row}'] = req_id
    ws2[f'B{row}'] = name
    ws2[f'C{row}'] = desc
    ws2[f'D{row}'] = priority

    for col in ['A', 'B', 'C', 'D']:
        set_cell_style(ws2[f'{col}{row}'], normal_font)

    if '\n' in desc and desc.count('\n') > 1:
        ws2.row_dimensions[row].height = 60
    elif '\n' in desc:
        ws2.row_dimensions[row].height = 40
    else:
        ws2.row_dimensions[row].height = 30
    row += 1

# ====================
# 3. 실력 지표 관리
# ====================
ws3 = wb.create_sheet("3.실력지표관리")
ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 75

# 제목
row = 1
ws3[f'A{row}'] = "실력 지표 관리 시스템"
ws3.merge_cells(f'A{row}:B{row}')
set_cell_style(ws3[f'A{row}'], Font(bold=True, size=14, color="FFFFFF"),
               purple_fill, Alignment(horizontal='center', vertical='center'))
ws3.row_dimensions[row].height = 30
row += 2

# 실력 점수
ws3[f'A{row}'] = "1. 실력 점수 (Skill Score)"
ws3.merge_cells(f'A{row}:B{row}')
set_cell_style(ws3[f'A{row}'], Font(bold=True, size=12), subheader_fill)
row += 1

score_info = [
    ("범위", "0~100 (Float)"),
    ("의미", "높을수록 초보, 낮을수록 실력자"),
    ("기본값", "50.0"),
    ("계산 방식", "최근 10개 문제 풀이 세션의 난이도 점수 평균"),
    ("자동 갱신", "auto 모드에서 문제 해결 시마다 자동 계산"),
]

for label, value in score_info:
    ws3[f'A{row}'] = label
    ws3[f'B{row}'] = value
    set_cell_style(ws3[f'A{row}'], bold_font)
    set_cell_style(ws3[f'B{row}'], normal_font)
    ws3.row_dimensions[row].height = 25
    row += 1

row += 1

# 실력 모드
ws3[f'A{row}'] = "2. 실력 모드 (Skill Mode)"
ws3.merge_cells(f'A{row}:B{row}')
set_cell_style(ws3[f'A{row}'], Font(bold=True, size=12), subheader_fill)
row += 1

ws3[f'A{row}'] = "자동 모드 (auto)"
ws3[f'B{row}'] = ("• 문제를 풀 때마다 실력 점수가 자동으로 업데이트됨\n"
                  "• 실력 점수에 따라 힌트 레벨이 자동 조정됨\n"
                  "• 관리자도 실력 점수와 힌트 레벨을 수정할 수 없음 (읽기 전용)")
set_cell_style(ws3[f'A{row}'], bold_font)
set_cell_style(ws3[f'B{row}'], normal_font)
ws3.row_dimensions[row].height = 60
row += 1

ws3[f'A{row}'] = "수동 모드 (manual)"
ws3[f'B{row}'] = ("• 관리자가 설정한 값으로 고정됨\n"
                  "• 문제를 풀어도 자동 업데이트되지 않음\n"
                  "• 관리자가 실력 점수와 힌트 레벨을 직접 수정 가능")
set_cell_style(ws3[f'A{row}'], bold_font)
set_cell_style(ws3[f'B{row}'], normal_font)
ws3.row_dimensions[row].height = 60
row += 1

row += 1

# 힌트 레벨
ws3[f'A{row}'] = "3. 힌트 레벨 (Hint Level)"
ws3.merge_cells(f'A{row}:B{row}')
set_cell_style(ws3[f'A{row}'], Font(bold=True, size=12), subheader_fill)
row += 1

hint_levels = [
    ("레벨 1 (기초)", "실력 점수 70 이상\n구체적 힌트 (함수명, 라이브러리 직접 언급)\n150자 이내"),
    ("레벨 2 (보통)", "실력 점수 40~70\n개념 힌트 (자료구조, 알고리즘 개념 안내)\n180자 이내"),
    ("레벨 3 (실력자)", "실력 점수 40 미만\n소크라테스식 질문 힌트\n200자 이내"),
]

for level, desc in hint_levels:
    ws3[f'A{row}'] = level
    ws3[f'B{row}'] = desc
    set_cell_style(ws3[f'A{row}'], bold_font)
    set_cell_style(ws3[f'B{row}'], normal_font)
    ws3.row_dimensions[row].height = 50
    row += 1

# ====================
# 4. 난이도 점수 계산
# ====================
ws4 = wb.create_sheet("4.난이도점수계산")
ws4.column_dimensions['A'].width = 25
ws4.column_dimensions['B'].width = 75

# 제목
row = 1
ws4[f'A{row}'] = "난이도 점수 계산 알고리즘"
ws4.merge_cells(f'A{row}:B{row}')
set_cell_style(ws4[f'A{row}'], Font(bold=True, size=14, color="FFFFFF"),
               purple_fill, Alignment(horizontal='center', vertical='center'))
ws4.row_dimensions[row].height = 30
row += 2

ws4[f'A{row}'] = "개요"
ws4[f'B{row}'] = "총점 = 0~100 (높을수록 사용자가 문제를 어려워함)"
set_cell_style(ws4[f'A{row}'], bold_font, subheader_fill)
set_cell_style(ws4[f'B{row}'], normal_font)
ws4.row_dimensions[row].height = 25
row += 2

# 계산 항목
ws4[f'A{row}'] = "1. 힌트 요청 시간 점수"
ws4[f'B{row}'] = ("• 1분 이내: +30점\n"
                  "• 5분 이내: +20점\n"
                  "• 10분 이내: +10점\n"
                  "• 10분 초과: 0점\n\n"
                  "빠르게 힌트를 요청할수록 문제를 어려워한다고 판단")
set_cell_style(ws4[f'A{row}'], bold_font)
set_cell_style(ws4[f'B{row}'], normal_font)
ws4.row_dimensions[row].height = 90
row += 1

ws4[f'A{row}'] = "2. 힌트 요청 횟수"
ws4[f'B{row}'] = ("• 계산식: 횟수 × 10점\n"
                  "• 최대값: 30점\n\n"
                  "힌트를 많이 요청할수록 문제를 어려워한다고 판단")
set_cell_style(ws4[f'A{row}'], bold_font)
set_cell_style(ws4[f'B{row}'], normal_font)
ws4.row_dimensions[row].height = 70
row += 1

ws4[f'A{row}'] = "3. 코드 실행 횟수"
ws4[f'B{row}'] = ("• 계산식: 횟수 × 3점\n"
                  "• 최대값: 20점\n\n"
                  "시행착오가 많을수록 문제를 어려워한다고 판단")
set_cell_style(ws4[f'A{row}'], bold_font)
set_cell_style(ws4[f'B{row}'], normal_font)
ws4.row_dimensions[row].height = 70
row += 1

ws4[f'A{row}'] = "4. 코드 길이"
ws4[f'B{row}'] = ("• 50자 미만: +20점\n"
                  "• 100자 미만: +10점\n"
                  "• 100자 이상: 0점\n\n"
                  "코드가 짧을수록 진전이 없다고 판단")
set_cell_style(ws4[f'A{row}'], bold_font)
set_cell_style(ws4[f'B{row}'], normal_font)
ws4.row_dimensions[row].height = 80
row += 2

ws4[f'A{row}'] = "최종 점수"
ws4[f'B{row}'] = "위 4가지 항목의 합계 (최대 100점)"
set_cell_style(ws4[f'A{row}'], bold_font, subheader_fill)
set_cell_style(ws4[f'B{row}'], normal_font)
ws4.row_dimensions[row].height = 25

# ====================
# 5. 레벨별 힌트 스타일
# ====================
ws5 = wb.create_sheet("5.레벨별힌트")
ws5.column_dimensions['A'].width = 30
ws5.column_dimensions['B'].width = 70

# 제목
row = 1
ws5[f'A{row}'] = "레벨별 힌트 스타일"
ws5.merge_cells(f'A{row}:B{row}')
set_cell_style(ws5[f'A{row}'], Font(bold=True, size=14, color="FFFFFF"),
               purple_fill, Alignment(horizontal='center', vertical='center'))
ws5.row_dimensions[row].height = 30
row += 2

# 레벨 1
ws5[f'A{row}'] = "레벨 1 (기초 - 구체적 힌트)"
ws5.merge_cells(f'A{row}:B{row}')
set_cell_style(ws5[f'A{row}'], Font(bold=True, size=12), subheader_fill)
row += 1

level1_data = [
    ("대상", "실력 점수 70 이상 (초보자)"),
    ("특징", "• 필요한 함수명, 라이브러리, 메서드를 직접 언급\n• 단계별 구체적인 코드 작성 방법 제시\n• 150자 이내로 간단명료하게 제공"),
    ("예시", '"N줄을 입력받으려면 for _ in range(N)을 사용하고, 각 줄을 list()로 변환해서 board에 append() 하세요."'),
]

for label, value in level1_data:
    ws5[f'A{row}'] = label
    ws5[f'B{row}'] = value
    set_cell_style(ws5[f'A{row}'], bold_font)
    set_cell_style(ws5[f'B{row}'], normal_font)
    if label == "특징":
        ws5.row_dimensions[row].height = 60
    elif label == "예시":
        ws5.row_dimensions[row].height = 50
    else:
        ws5.row_dimensions[row].height = 25
    row += 1

row += 1

# 레벨 2
ws5[f'A{row}'] = "레벨 2 (보통 - 개념 힌트)"
ws5.merge_cells(f'A{row}:B{row}')
set_cell_style(ws5[f'A{row}'], Font(bold=True, size=12), subheader_fill)
row += 1

level2_data = [
    ("대상", "실력 점수 40~70 (중급자)"),
    ("특징", "• 함수명을 직접 언급하지 않고 개념으로 유도\n• 필요한 자료구조나 알고리즘 개념 안내\n• 180자 이내로 제공"),
    ("예시", '"N줄의 보드를 저장하려면 2차원 리스트가 필요합니다. 반복문으로 각 줄을 입력받아 추가하세요."'),
]

for label, value in level2_data:
    ws5[f'A{row}'] = label
    ws5[f'B{row}'] = value
    set_cell_style(ws5[f'A{row}'], bold_font)
    set_cell_style(ws5[f'B{row}'], normal_font)
    if label == "특징":
        ws5.row_dimensions[row].height = 60
    elif label == "예시":
        ws5.row_dimensions[row].height = 40
    else:
        ws5.row_dimensions[row].height = 25
    row += 1

row += 1

# 레벨 3
ws5[f'A{row}'] = "레벨 3 (실력자 - 소크라테스식)"
ws5.merge_cells(f'A{row}:B{row}')
set_cell_style(ws5[f'A{row}'], Font(bold=True, size=12), subheader_fill)
row += 1

level3_data = [
    ("대상", "실력 점수 40 미만 (고급자)"),
    ("특징", "• 질문 형식으로만 힌트 제공\n• 학생이 스스로 답을 찾도록 유도\n• 한 번에 하나의 질문만 (200자 이내)\n• 평가나 분석 내용 절대 포함하지 않음"),
    ("예시 (올바름)", '"전체 보드의 상태를 어떻게 입력받아 저장할 수 있을까요?"'),
    ("예시 (잘못됨)", '"학생이 입력을 받는 부분까지 올바르게 작성했습니다. 이제 다음 단계는 보드 상태를 받아와야 합니다."'),
]

for label, value in level3_data:
    ws5[f'A{row}'] = label
    ws5[f'B{row}'] = value
    set_cell_style(ws5[f'A{row}'], bold_font)
    set_cell_style(ws5[f'B{row}'], normal_font)
    if label == "특징":
        ws5.row_dimensions[row].height = 70
    else:
        ws5.row_dimensions[row].height = 35
    row += 1

# ====================
# 6. 데이터 모델
# ====================
ws6 = wb.create_sheet("6.데이터모델")
ws6.column_dimensions['A'].width = 20
ws6.column_dimensions['B'].width = 20
ws6.column_dimensions['C'].width = 15
ws6.column_dimensions['D'].width = 50

# 제목
row = 1
ws6[f'A{row}'] = "데이터 모델 정의"
ws6.merge_cells(f'A{row}:D{row}')
set_cell_style(ws6[f'A{row}'], Font(bold=True, size=14, color="FFFFFF"),
               purple_fill, Alignment(horizontal='center', vertical='center'))
ws6.row_dimensions[row].height = 30
row += 2

# User 모델
ws6[f'A{row}'] = "User 모델 (확장)"
ws6.merge_cells(f'A{row}:D{row}')
set_cell_style(ws6[f'A{row}'], white_font, header_fill, Alignment(horizontal='center'))
row += 1

headers = ["필드명", "타입", "기본값", "설명"]
for col, header in enumerate(headers, 1):
    cell = ws6.cell(row, col, header)
    set_cell_style(cell, bold_font, subheader_fill, Alignment(horizontal='center'))
row += 1

user_fields = [
    ("skill_score", "FloatField", "50.0", "실력 점수 (0~100, 높을수록 초보)"),
    ("skill_mode", "CharField", "'auto'", "실력 지표 모드 ('auto' | 'manual')"),
    ("hint_level", "IntegerField", "2", "힌트 레벨 (1:기초, 2:보통, 3:실력자)"),
]

for field_name, field_type, default, desc in user_fields:
    ws6[f'A{row}'] = field_name
    ws6[f'B{row}'] = field_type
    ws6[f'C{row}'] = default
    ws6[f'D{row}'] = desc
    for col in ['A', 'B', 'C', 'D']:
        set_cell_style(ws6[f'{col}{row}'], normal_font)
    ws6.row_dimensions[row].height = 25
    row += 1

row += 2

# ProblemSession 모델
ws6[f'A{row}'] = "ProblemSession 모델 (신규)"
ws6.merge_cells(f'A{row}:D{row}')
set_cell_style(ws6[f'A{row}'], white_font, header_fill, Alignment(horizontal='center'))
row += 1

for col, header in enumerate(headers, 1):
    cell = ws6.cell(row, col, header)
    set_cell_style(cell, bold_font, subheader_fill, Alignment(horizontal='center'))
row += 1

session_fields = [
    ("user", "ForeignKey", "-", "사용자 (User 모델 참조)"),
    ("problem", "ForeignKey", "-", "문제 (Problem 모델 참조)"),
    ("started_at", "DateTimeField", "auto_now_add", "세션 시작 시간"),
    ("first_hint_at", "DateTimeField", "null", "첫 힌트 요청 시간"),
    ("first_run_at", "DateTimeField", "null", "첫 코드 실행 시간"),
    ("solved_at", "DateTimeField", "null", "문제 해결 시간"),
    ("hint_count", "IntegerField", "0", "힌트 요청 횟수"),
    ("run_count", "IntegerField", "0", "코드 실행 횟수"),
    ("max_code_length", "IntegerField", "0", "최대 코드 길이"),
    ("is_solved", "BooleanField", "False", "해결 여부"),
]

for field_name, field_type, default, desc in session_fields:
    ws6[f'A{row}'] = field_name
    ws6[f'B{row}'] = field_type
    ws6[f'C{row}'] = default
    ws6[f'D{row}'] = desc
    for col in ['A', 'B', 'C', 'D']:
        set_cell_style(ws6[f'{col}{row}'], normal_font)
    ws6.row_dimensions[row].height = 25
    row += 1

row += 2
ws6[f'A{row}'] = "제약사항"
ws6[f'B{row}'] = "unique_together = ['user', 'problem']"
ws6.merge_cells(f'B{row}:D{row}')
set_cell_style(ws6[f'A{row}'], bold_font, subheader_fill)
set_cell_style(ws6[f'B{row}'], normal_font)
ws6.row_dimensions[row].height = 25

# ====================
# 7. API 명세
# ====================
ws7 = wb.create_sheet("7.API명세")
ws7.column_dimensions['A'].width = 15
ws7.column_dimensions['B'].width = 40
ws7.column_dimensions['C'].width = 15
ws7.column_dimensions['D'].width = 35

# 제목
row = 1
ws7[f'A{row}'] = "API 명세"
ws7.merge_cells(f'A{row}:D{row}')
set_cell_style(ws7[f'A{row}'], Font(bold=True, size=14, color="FFFFFF"),
               purple_fill, Alignment(horizontal='center', vertical='center'))
ws7.row_dimensions[row].height = 30
row += 2

# 헤더
headers = ["구분", "엔드포인트", "메서드", "설명"]
for col, header in enumerate(headers, 1):
    cell = ws7.cell(row, col, header)
    set_cell_style(cell, white_font, header_fill, Alignment(horizontal='center'))
row += 1

apis = [
    ("힌트 요청", "/api/v1/coding-test/hints/", "POST", "사용자 힌트 레벨에 따른 AI 힌트 생성"),
    ("사용자 목록", "/api/v1/admin/users/", "GET", "모든 사용자 목록 조회 (실력 지표 포함)"),
    ("실력 지표 수정", "/api/v1/admin/users/{id}/skill/", "PATCH", "사용자 실력 지표 업데이트"),
    ("사용자 상태", "/api/v1/admin/users/{id}/", "PATCH", "사용자 활성화 상태 변경"),
    ("사용자 삭제", "/api/v1/admin/users/{id}/delete/", "DELETE", "사용자 삭제"),
    ("AI 설정 조회", "/api/v1/coding-test/ai-config/", "GET", "AI 모델 설정 조회"),
    ("AI 설정 업데이트", "/api/v1/coding-test/ai-config/update/", "POST", "AI 모델 설정 업데이트"),
]

for category, endpoint, method, desc in apis:
    ws7[f'A{row}'] = category
    ws7[f'B{row}'] = endpoint
    ws7[f'C{row}'] = method
    ws7[f'D{row}'] = desc
    for col in ['A', 'B', 'C', 'D']:
        set_cell_style(ws7[f'{col}{row}'], normal_font)
    ws7.row_dimensions[row].height = 30
    row += 1

# ====================
# 8. 시스템 플로우
# ====================
ws8 = wb.create_sheet("8.시스템플로우")
ws8.column_dimensions['A'].width = 30
ws8.column_dimensions['B'].width = 70

# 제목
row = 1
ws8[f'A{row}'] = "시스템 플로우"
ws8.merge_cells(f'A{row}:B{row}')
set_cell_style(ws8[f'A{row}'], Font(bold=True, size=14, color="FFFFFF"),
               purple_fill, Alignment(horizontal='center', vertical='center'))
ws8.row_dimensions[row].height = 30
row += 2

# 자동 모드 플로우
ws8[f'A{row}'] = "자동 모드 - 실력 지표 갱신"
ws8.merge_cells(f'A{row}:B{row}')
set_cell_style(ws8[f'A{row}'], Font(bold=True, size=12), subheader_fill)
row += 1

flow_auto = """1. 사용자가 문제 풀이 시작
   ↓
2. ProblemSession 생성 (started_at 기록)
   ↓
3. 힌트 요청 시
   - first_hint_at 기록 (첫 요청인 경우)
   - hint_count 증가
   ↓
4. 코드 실행 시
   - first_run_at 기록 (첫 실행인 경우)
   - run_count 증가
   - max_code_length 업데이트
   ↓
5. 문제 해결 시
   - solved_at 기록
   - is_solved = True
   - user.update_skill_score() 호출
   ↓
6. update_skill_score() 메서드
   - skill_mode가 'auto'인지 확인
   - 최근 10개 세션의 난이도 점수 계산
   - 평균 점수로 skill_score 업데이트
   - 점수에 따라 hint_level 자동 조정
   ↓
7. 다음 힌트 요청 시 새로운 hint_level 적용"""

ws8[f'A{row}'] = flow_auto
ws8.merge_cells(f'A{row}:B{row}')
set_cell_style(ws8[f'A{row}'], normal_font)
ws8.row_dimensions[row].height = 300
row += 2

# 수동 모드 플로우
ws8[f'A{row}'] = "수동 모드 - 관리자 설정"
ws8.merge_cells(f'A{row}:B{row}')
set_cell_style(ws8[f'A{row}'], Font(bold=True, size=12), subheader_fill)
row += 1

flow_manual = """1. 관리자가 "관리자 패널" 접속
   ↓
2. "사용자 관리" 탭 선택
   ↓
3. 특정 사용자의 "실력 수정" 클릭
   ↓
4. 편집 모드 진입
   - skill_mode를 'manual'로 선택
   - skill_score 직접 입력 (예: 75.5)
   - hint_level 직접 선택 (예: 레벨 1)
   ↓
5. "저장" 버튼 클릭
   ↓
6. API 호출: PATCH /api/v1/admin/users/{id}/skill/
   ↓
7. 백엔드: User 모델 업데이트
   - skill_score = 75.5
   - skill_mode = 'manual'
   - hint_level = 1
   ↓
8. 사용자는 다음 힌트 요청부터 레벨 1 힌트 받음
   (문제를 풀어도 자동 업데이트되지 않음)"""

ws8[f'A{row}'] = flow_manual
ws8.merge_cells(f'A{row}:B{row}')
set_cell_style(ws8[f'A{row}'], normal_font)
ws8.row_dimensions[row].height = 260

# ====================
# 9. 테스트 시나리오
# ====================
ws9 = wb.create_sheet("9.테스트시나리오")
ws9.column_dimensions['A'].width = 12
ws9.column_dimensions['B'].width = 28
ws9.column_dimensions['C'].width = 45
ws9.column_dimensions['D'].width = 25

# 제목
row = 1
ws9[f'A{row}'] = "테스트 시나리오"
ws9.merge_cells(f'A{row}:D{row}')
set_cell_style(ws9[f'A{row}'], Font(bold=True, size=14, color="FFFFFF"),
               purple_fill, Alignment(horizontal='center', vertical='center'))
ws9.row_dimensions[row].height = 30
row += 2

# 헤더
headers = ["ID", "테스트 항목", "테스트 절차", "예상 결과"]
for col, header in enumerate(headers, 1):
    cell = ws9.cell(row, col, header)
    set_cell_style(cell, white_font, header_fill, Alignment(horizontal='center'))
row += 1

tests = [
    ("TC-001", "힌트 레벨별 응답 확인",
     "1. 사용자 hint_level을 1, 2, 3으로 각각 설정\n2. 동일 문제에 대해 힌트 요청\n3. 응답 내용 비교",
     "레벨별로 다른 형태의 힌트 제공됨"),

    ("TC-002", "실력 점수 자동 계산",
     "1. 새 사용자 생성 (auto 모드)\n2. 여러 문제 풀이 (힌트/실행 횟수 다양하게)\n3. skill_score 확인",
     "행동 패턴에 따라 점수가 계산됨"),

    ("TC-003", "힌트 레벨 자동 조정",
     "1. 사용자 skill_score를 75로 설정\n2. update_skill_score() 호출\n3. hint_level 확인",
     "hint_level이 1로 자동 변경됨"),

    ("TC-004", "manual 모드 고정 확인",
     "1. 사용자를 manual 모드로 설정\n2. skill_score=30, hint_level=1로 고정\n3. 문제 풀이 후 확인",
     "값이 변경되지 않고 유지됨"),

    ("TC-005", "관리자 실력 지표 수정",
     "1. 관리자로 로그인\n2. UsersTab에서 사용자 실력 수정\n3. API 호출 확인",
     "변경사항이 DB에 반영됨"),
]

for test_id, test_name, procedure, expected in tests:
    ws9[f'A{row}'] = test_id
    ws9[f'B{row}'] = test_name
    ws9[f'C{row}'] = procedure
    ws9[f'D{row}'] = expected

    for col in ['A', 'B', 'C', 'D']:
        set_cell_style(ws9[f'{col}{row}'], normal_font)

    ws9.row_dimensions[row].height = 70
    row += 1

# ====================
# 10. 제약사항 및 향후 개선
# ====================
ws10 = wb.create_sheet("10.제약사항및개선")
ws10.column_dimensions['A'].width = 20
ws10.column_dimensions['B'].width = 80

# 제목
row = 1
ws10[f'A{row}'] = "제약사항 및 향후 개선사항"
ws10.merge_cells(f'A{row}:B{row}')
set_cell_style(ws10[f'A{row}'], Font(bold=True, size=14, color="FFFFFF"),
               purple_fill, Alignment(horizontal='center', vertical='center'))
ws10.row_dimensions[row].height = 30
row += 2

# 제약사항
ws10[f'A{row}'] = "제약사항"
ws10.merge_cells(f'A{row}:B{row}')
set_cell_style(ws10[f'A{row}'], Font(bold=True, size=12), subheader_fill)
row += 1

constraints = [
    ("Hugging Face API", "월 30,000회 무료 제한, 0.5B/1.5B 모델은 Inference Providers 미지원"),
    ("세션 추적", "현재는 모델만 정의됨, 실제 추적 로직은 향후 구현 예정"),
    ("권한", "일반 사용자는 자신의 실력 지표를 볼 수 없음 (관리자만 조회 가능)"),
    ("계산 기준", "실력 점수는 최근 10개 세션만 고려"),
]

for label, desc in constraints:
    ws10[f'A{row}'] = label
    ws10[f'B{row}'] = desc
    set_cell_style(ws10[f'A{row}'], bold_font)
    set_cell_style(ws10[f'B{row}'], normal_font)
    ws10.row_dimensions[row].height = 40
    row += 1

row += 1

# 향후 개선사항
ws10[f'A{row}'] = "향후 개선사항"
ws10.merge_cells(f'A{row}:B{row}')
set_cell_style(ws10[f'A{row}'], Font(bold=True, size=12), subheader_fill)
row += 1

improvements = [
    ("실시간 세션 추적", "문제 선택 시 ProblemSession 자동 생성, 힌트/실행 버튼 클릭 시 자동 기록"),
    ("대시보드 추가", "사용자별 실력 점수 변화 그래프, 힌트 레벨 분포 차트, 평균 해결 시간 통계"),
    ("알고리즘 개선", "문제 난이도별 가중치 적용, 시간대별 학습 패턴 분석, 더 세분화된 힌트 레벨 (5단계 등)"),
    ("개인화 학습 경로", "실력 점수 기반 추천 문제 알고리즘 개발"),
    ("힌트 효과 분석", "힌트 제공 후 문제 해결율 통계 분석 기능"),
]

for label, desc in improvements:
    ws10[f'A{row}'] = label
    ws10[f'B{row}'] = desc
    set_cell_style(ws10[f'A{row}'], bold_font)
    set_cell_style(ws10[f'B{row}'], normal_font)
    ws10.row_dimensions[row].height = 40
    row += 1

# 파일 저장
output_path = 'C:/Users/playdata2/Desktop/playdata/Workspace/팀프로젝트5/5th-project_mvp/docs/개인화힌트시스템_요구사항정의서_v1.0.xlsx'
wb.save(output_path)
print("Requirements document created successfully:", output_path)
print("\nSheets created:")
for sheet in wb.sheetnames:
    print(f"  - {sheet}")
