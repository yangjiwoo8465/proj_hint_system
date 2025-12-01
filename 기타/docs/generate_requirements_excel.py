"""
요구사항 정의서 엑셀 파일 생성
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 워크북 생성
wb = Workbook()
wb.remove(wb.active)  # 기본 시트 제거

# 스타일 정의
header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
subheader_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
white_font = Font(color="FFFFFF", bold=True, size=12)
bold_font = Font(bold=True, size=11)
normal_font = Font(size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def set_cell_style(cell, font, fill=None, alignment=None, border=border):
    cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    else:
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    cell.border = border

# ====================
# 1. 프로젝트 개요
# ====================
ws1 = wb.create_sheet("1. 프로젝트 개요")
ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 80

data = [
    ("프로젝트명", "개인화된 힌트 시스템 (Personalized Hint System)"),
    ("작성일", "2025-01-06"),
    ("작성자", "개발팀"),
    ("프로젝트 목적", "사용자의 실력 수준에 맞는 개인화된 힌트를 제공하여 효과적인 학습을 지원"),
    ("주요 기능", "• 사용자 행동 추적 및 실력 점수 자동 계산\n• 3단계 힌트 레벨 시스템 (기초/보통/실력자)\n• 관리자 패널을 통한 사용자 실력 지표 관리\n• 자동/수동 모드 전환"),
    ("기술 스택", "• Frontend: React 18, Redux, Monaco Editor\n• Backend: Django 4.2, Django REST Framework\n• Database: PostgreSQL\n• AI: Hugging Face Inference Providers (Qwen 7B/32B)\n• Infrastructure: Docker, Docker Compose"),
]

row = 1
for label, value in data:
    ws1[f'A{row}'] = label
    ws1[f'B{row}'] = value
    set_cell_style(ws1[f'A{row}'], bold_font, subheader_fill)
    set_cell_style(ws1[f'B{row}'], normal_font)
    ws1.row_dimensions[row].height = 60 if label in ["주요 기능", "기술 스택"] else None
    row += 1

# ====================
# 2. 기능 요구사항
# ====================
ws2 = wb.create_sheet("2. 기능 요구사항")
ws2.column_dimensions['A'].width = 10
ws2.column_dimensions['B'].width = 25
ws2.column_dimensions['C'].width = 60
ws2.column_dimensions['D'].width = 15

# 헤더
headers = ["ID", "기능명", "설명", "우선순위"]
for col, header in enumerate(headers, 1):
    cell = ws2.cell(1, col, header)
    set_cell_style(cell, white_font, header_fill, Alignment(horizontal='center', vertical='center'))

# 데이터
requirements = [
    ("FR-001", "사용자 행동 추적", "문제 선택 시점부터 첫 힌트 요청까지의 시간, 첫 실행까지의 시간, 실행 횟수, 코드 길이 등을 추적", "높음"),
    ("FR-002", "실력 점수 자동 계산", "ProblemSession 데이터를 기반으로 0-100 점수 자동 계산 (높을수록 초보)", "높음"),
    ("FR-003", "힌트 레벨 자동 조정", "실력 점수에 따라 힌트 레벨을 자동 조정 (70+ = 레벨1, 40-70 = 레벨2, 40- = 레벨3)", "높음"),
    ("FR-004", "레벨별 힌트 제공", "레벨 1: 구체적 힌트 (함수명 직접 언급)\n레벨 2: 개념 힌트 (개념/자료구조 안내)\n레벨 3: 소크라테스식 질문", "높음"),
    ("FR-005", "관리자 사용자 목록 조회", "관리자가 모든 사용자의 실력 지표를 조회", "중간"),
    ("FR-006", "사용자 실력 지표 수정", "관리자가 skill_score, skill_mode, hint_level 수정 가능", "중간"),
    ("FR-007", "자동/수동 모드 전환", "auto 모드: 자동 계산 및 조정\nmanual 모드: 관리자 설정값 고정", "중간"),
    ("FR-008", "사용자 활성화/비활성화", "관리자가 사용자 계정 활성 상태 관리", "낮음"),
    ("FR-009", "사용자 삭제", "관리자가 사용자 계정 삭제", "낮음"),
]

row = 2
for req_id, name, desc, priority in requirements:
    ws2[f'A{row}'] = req_id
    ws2[f'B{row}'] = name
    ws2[f'C{row}'] = desc
    ws2[f'D{row}'] = priority

    for col in ['A', 'B', 'C', 'D']:
        set_cell_style(ws2[f'{col}{row}'], normal_font)

    ws2.row_dimensions[row].height = 50 if '\n' in desc else None
    row += 1

# ====================
# 3. 데이터 모델
# ====================
ws3 = wb.create_sheet("3. 데이터 모델")
ws3.column_dimensions['A'].width = 15
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 15
ws3.column_dimensions['D'].width = 50

# User 모델
row = 1
ws3[f'A{row}'] = "User (확장)"
ws3.merge_cells(f'A{row}:D{row}')
set_cell_style(ws3[f'A{row}'], white_font, header_fill, Alignment(horizontal='center'))
row += 1

user_fields = [
    ("필드명", "타입", "기본값", "설명"),
    ("skill_score", "FloatField", "50.0", "실력 점수 (0~100, 높을수록 초보)"),
    ("skill_mode", "CharField", "auto", "실력 지표 모드 (auto/manual)"),
    ("hint_level", "IntegerField", "2", "힌트 레벨 (1:기초, 2:보통, 3:실력자)"),
]

for field in user_fields:
    for col, value in enumerate(field, 1):
        cell = ws3.cell(row, col, value)
        if row == 2:
            set_cell_style(cell, bold_font, subheader_fill, Alignment(horizontal='center'))
        else:
            set_cell_style(cell, normal_font)
    row += 1

row += 1

# ProblemSession 모델
ws3[f'A{row}'] = "ProblemSession (신규)"
ws3.merge_cells(f'A{row}:D{row}')
set_cell_style(ws3[f'A{row}'], white_font, header_fill, Alignment(horizontal='center'))
row += 1

session_fields = [
    ("필드명", "타입", "기본값", "설명"),
    ("user", "ForeignKey", "-", "사용자 (User 모델 참조)"),
    ("problem", "ForeignKey", "-", "문제 (Problem 모델 참조)"),
    ("started_at", "DateTimeField", "auto_now_add", "세션 시작 시간"),
    ("first_hint_at", "DateTimeField", "null", "첫 힌트 요청 시간"),
    ("first_run_at", "DateTimeField", "null", "첫 코드 실행 시간"),
    ("solved_at", "DateTimeField", "null", "문제 해결 시간"),
    ("hint_count", "IntegerField", "0", "힌트 요청 횟수"),
    ("run_count", "IntegerField", "0", "코드 실행 횟수"),
    ("max_code_length", "IntegerField", "0", "최대 코드 길이"),
    ("is_solved", "BooleanField", "False", "해결 여부"),
]

for field in session_fields:
    for col, value in enumerate(field, 1):
        cell = ws3.cell(row, col, value)
        if row == (len(user_fields) + 3):
            set_cell_style(cell, bold_font, subheader_fill, Alignment(horizontal='center'))
        else:
            set_cell_style(cell, normal_font)
    row += 1

# ====================
# 4. API 명세
# ====================
ws4 = wb.create_sheet("4. API 명세")
ws4.column_dimensions['A'].width = 10
ws4.column_dimensions['B'].width = 40
ws4.column_dimensions['C'].width = 15
ws4.column_dimensions['D'].width = 45

# 헤더
headers = ["구분", "엔드포인트", "메서드", "설명"]
for col, header in enumerate(headers, 1):
    cell = ws4.cell(1, col, header)
    set_cell_style(cell, white_font, header_fill, Alignment(horizontal='center', vertical='center'))

apis = [
    ("힌트", "/coding-test/hint/", "POST", "사용자 힌트 레벨에 따른 AI 힌트 생성"),
    ("관리자", "/admin/users/", "GET", "모든 사용자 목록 조회 (실력 지표 포함)"),
    ("관리자", "/admin/users/{id}/", "PATCH", "사용자 활성화 상태 변경"),
    ("관리자", "/admin/users/{id}/skill/", "PATCH", "사용자 실력 지표 업데이트"),
    ("관리자", "/admin/users/{id}/delete/", "DELETE", "사용자 삭제"),
]

row = 2
for category, endpoint, method, desc in apis:
    ws4[f'A{row}'] = category
    ws4[f'B{row}'] = endpoint
    ws4[f'C{row}'] = method
    ws4[f'D{row}'] = desc

    for col in ['A', 'B', 'C', 'D']:
        set_cell_style(ws4[f'{col}{row}'], normal_font)

    row += 1

# ====================
# 5. 비즈니스 로직
# ====================
ws5 = wb.create_sheet("5. 비즈니스 로직")
ws5.column_dimensions['A'].width = 25
ws5.column_dimensions['B'].width = 75

logics = [
    ("실력 점수 계산 알고리즘",
     "1. 힌트 요청 시간: 1분 이내(+30), 5분 이내(+20), 10분 이내(+10)\n"
     "2. 힌트 요청 횟수: 횟수 × 10 (최대 30점)\n"
     "3. 실행 횟수: 횟수 × 3 (최대 20점)\n"
     "4. 코드 길이: 50자 미만(+20), 100자 미만(+10)\n"
     "5. 최종 점수는 100점을 초과하지 않음"),

    ("힌트 레벨 자동 조정 규칙",
     "• 최근 10개 세션의 평균 점수 계산\n"
     "• 70점 이상 → 레벨 1 (기초)\n"
     "• 40~70점 → 레벨 2 (보통)\n"
     "• 40점 미만 → 레벨 3 (실력자)\n"
     "• auto 모드에서만 자동 조정됨"),

    ("레벨별 힌트 프롬프트",
     "레벨 1 (기초):\n"
     "• 함수명, 라이브러리를 직접 언급\n"
     "• 예: 'for _ in range(N)을 사용하세요'\n\n"
     "레벨 2 (보통):\n"
     "• 개념과 자료구조 위주\n"
     "• 예: '2차원 리스트가 필요합니다'\n\n"
     "레벨 3 (실력자):\n"
     "• 소크라테스식 질문\n"
     "• 예: '어떻게 저장할 수 있을까요?'"),

    ("자동/수동 모드 동작",
     "auto 모드:\n"
     "• 문제 해결 후 update_skill_score() 자동 호출\n"
     "• skill_score 및 hint_level 자동 갱신\n\n"
     "manual 모드:\n"
     "• 관리자가 설정한 값으로 고정\n"
     "• update_skill_score() 호출되지 않음"),
]

row = 1
for logic_name, logic_desc in logics:
    ws5[f'A{row}'] = logic_name
    ws5[f'B{row}'] = logic_desc
    set_cell_style(ws5[f'A{row}'], bold_font, subheader_fill)
    set_cell_style(ws5[f'B{row}'], normal_font)
    ws5.row_dimensions[row].height = 120 if "프롬프트" in logic_name else 80
    row += 1

# ====================
# 6. 시스템 플로우
# ====================
ws6 = wb.create_sheet("6. 시스템 플로우")
ws6.column_dimensions['A'].width = 30
ws6.column_dimensions['B'].width = 70

flows = [
    ("1. 힌트 요청 플로우",
     "① 사용자가 'CodingTest' 페이지에서 문제 선택\n"
     "② 힌트 버튼 클릭 시 POST /coding-test/hint/ 호출\n"
     "③ Backend에서 user.hint_level 조회\n"
     "④ 레벨에 맞는 프롬프트 생성\n"
     "⑤ Hugging Face API 호출 (Qwen 모델)\n"
     "⑥ AI 응답을 프론트엔드로 반환\n"
     "⑦ 사용자에게 힌트 표시"),

    ("2. 실력 점수 자동 갱신 플로우",
     "① 사용자가 문제 해결 완료\n"
     "② Backend에서 ProblemSession 생성/업데이트\n"
     "③ user.skill_mode == 'auto'인 경우\n"
     "④ user.update_skill_score() 호출\n"
     "⑤ 최근 10개 세션의 calculate_difficulty_score() 계산\n"
     "⑥ 평균 점수로 skill_score 업데이트\n"
     "⑦ 점수 구간에 따라 hint_level 자동 조정"),

    ("3. 관리자 실력 지표 수정 플로우",
     "① 관리자가 'AdminPanel - 사용자 관리' 접속\n"
     "② 사용자 목록에서 '실력 수정' 버튼 클릭\n"
     "③ inline editing 모드로 전환\n"
     "④ skill_score, skill_mode, hint_level 수정\n"
     "⑤ '저장' 버튼 클릭\n"
     "⑥ PATCH /admin/users/{id}/skill/ 호출\n"
     "⑦ Backend에서 유효성 검증 후 저장\n"
     "⑧ 업데이트된 데이터 반환 및 UI 갱신"),
]

row = 1
for flow_name, flow_desc in flows:
    ws6[f'A{row}'] = flow_name
    ws6[f'B{row}'] = flow_desc
    set_cell_style(ws6[f'A{row}'], bold_font, subheader_fill)
    set_cell_style(ws6[f'B{row}'], normal_font)
    ws6.row_dimensions[row].height = 120
    row += 1

# ====================
# 7. 테스트 시나리오
# ====================
ws7 = wb.create_sheet("7. 테스트 시나리오")
ws7.column_dimensions['A'].width = 10
ws7.column_dimensions['B'].width = 30
ws7.column_dimensions['C'].width = 50
ws7.column_dimensions['D'].width = 20

# 헤더
headers = ["ID", "테스트 항목", "테스트 절차", "예상 결과"]
for col, header in enumerate(headers, 1):
    cell = ws7.cell(1, col, header)
    set_cell_style(cell, white_font, header_fill, Alignment(horizontal='center', vertical='center'))

tests = [
    ("TC-001", "힌트 레벨별 응답 확인",
     "1. 사용자 hint_level을 1, 2, 3으로 각각 설정\n2. 동일 문제에 대해 힌트 요청\n3. 응답 내용 확인",
     "레벨별로 다른 형태의 힌트 제공됨"),

    ("TC-002", "실력 점수 자동 계산",
     "1. 새 사용자 생성 (auto 모드)\n2. 여러 문제 풀이 (힌트/실행 횟수 다양하게)\n3. skill_score 확인",
     "행동 패턴에 따라 점수가 계산됨"),

    ("TC-003", "힌트 레벨 자동 조정",
     "1. 사용자 skill_score를 75로 설정\n2. update_skill_score() 호출\n3. hint_level 확인",
     "hint_level이 1로 자동 변경됨"),

    ("TC-004", "manual 모드 고정 확인",
     "1. 사용자를 manual 모드로 설정\n2. skill_score=30, hint_level=1로 고정\n3. 문제 풀이 후 확인",
     "값이 변경되지 않고 유지됨"),

    ("TC-005", "관리자 실력 지표 수정",
     "1. 관리자로 로그인\n2. UsersTab에서 사용자 실력 수정\n3. API 호출 확인",
     "변경사항이 DB에 반영됨"),
]

row = 2
for test_id, test_name, procedure, expected in tests:
    ws7[f'A{row}'] = test_id
    ws7[f'B{row}'] = test_name
    ws7[f'C{row}'] = procedure
    ws7[f'D{row}'] = expected

    for col in ['A', 'B', 'C', 'D']:
        set_cell_style(ws7[f'{col}{row}'], normal_font)

    ws7.row_dimensions[row].height = 60
    row += 1

# ====================
# 8. 향후 개선사항
# ====================
ws8 = wb.create_sheet("8. 향후 개선사항")
ws8.column_dimensions['A'].width = 15
ws8.column_dimensions['B'].width = 85

improvements = [
    ("실시간 세션 추적", "현재는 모델만 정의됨. 실제 문제 풀이 중 행동 추적 로직 구현 필요"),
    ("대시보드 시각화", "관리자 패널에 실력 점수 분포, 힌트 레벨 통계 그래프 추가"),
    ("개인화 학습 경로", "실력 점수 기반 추천 문제 알고리즘 개발"),
    ("힌트 효과 분석", "힌트 제공 후 문제 해결율 통계 분석 기능"),
    ("다국어 힌트 지원", "영어, 일본어 등 다국어 힌트 프롬프트 추가"),
]

row = 1
for improvement, desc in improvements:
    ws8[f'A{row}'] = improvement
    ws8[f'B{row}'] = desc
    set_cell_style(ws8[f'A{row}'], bold_font, subheader_fill)
    set_cell_style(ws8[f'B{row}'], normal_font)
    ws8.row_dimensions[row].height = 40
    row += 1

# 파일 저장
output_path = 'C:/Users/playdata2/Desktop/playdata/Workspace/팀프로젝트5/5th-project_mvp/docs/요구사항_정의서.xlsx'
wb.save(output_path)
print("Excel file created successfully:", output_path)
